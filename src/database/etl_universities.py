
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import os
from dotenv import load_dotenv
import psycopg2
from openpyxl import load_workbook

load_dotenv()

DB = dict(host=os.getenv("DB_HOST", "127.0.0.1"), port=int(os.getenv("DB_PORT", 5432)), dbname=os.getenv("DB_NAME", "ownway_db"),
          user=os.getenv("DB_USER"), password=os.getenv("DB_PASS"))

CITY_MAP = {
    "Ankara": 1, "Antalya": 2, "Erzurum": 3,
    "Konya": 4, "Zonguldak": 5, "İstanbul": 6, "İzmir": 7,
}


def load_universities():
    ws = load_workbook("Universite.xlsx", data_only=True).active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        sehir     = str(r[0]).strip()
        uni_name  = str(r[1]).strip()
        uni_type  = str(r[2]).strip().capitalize()  
        has_campus = str(r[3]).strip().lower() == "evet"
        campus_count = int(r[4]) if r[4] else 0

        city_id = CITY_MAP.get(sehir)
        if not city_id:
            print(f"  [UYARI] Bilinmeyen sehir: '{sehir}' — atlanıyor.")
            continue

        rows.append((city_id, uni_name, uni_type, has_campus, campus_count))
    return rows


def db_write(rows):
    with psycopg2.connect(**DB) as conn, conn.cursor() as cur:
     
        cur.execute("DELETE FROM universities;")

        cur.executemany(
            "INSERT INTO universities (city_id, uni_name, uni_type, has_campus, campus_count)"
            " VALUES (%s, %s, %s, %s, %s);",
            rows,
        )

       
        cur.execute("""
            SELECT u.id, c.city_name, u.uni_name, u.uni_type, u.has_campus, u.campus_count
            FROM universities u
            JOIN cities c ON c.id = u.city_id
            ORDER BY c.city_name, u.uni_name;
        """)
        print(f"\n  {'id':>3}  {'Şehir':<12}  {'Üniversite':<35}  {'Tür':<8}  {'Kampüs':>7}  {'Sayı':>5}")
        print("  " + "─" * 76)
        for r in cur.fetchall():
            kampus = "Evet" if r[4] else "Hayır"
            print(f"  {r[0]:>3}  {r[1]:<12}  {r[2]:<35}  {r[3]:<8}  {kampus:>7}  {r[5]:>5}")


if __name__ == "__main__":
    rows = load_universities()
    print(f"  {len(rows)} üniversite yüklendi.\n")
    db_write(rows)
    print(f"\n  Universities tablosu güncellendi. ({len(rows)} kayıt)")
