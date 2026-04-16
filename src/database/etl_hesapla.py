
import os
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from dotenv import load_dotenv
import json
import psycopg2
from openpyxl import load_workbook

load_dotenv()

DB = dict(host=os.getenv("DB_HOST", "127.0.0.1"), port=int(os.getenv("DB_PORT", 5432)), dbname=os.getenv("DB_NAME", "ownway_db"),
          user=os.getenv("DB_USER"), password=os.getenv("DB_PASS"))

CITIES = ["İstanbul", "İzmir", "Ankara", "Konya", "Antalya", "Zonguldak", "Erzurum"]

MAP_MALIYET  = {"TR34 (İstanbul)": "İstanbul", "TR35 (İzmir)": "İzmir",
                "TR06 (Ankara)": "Ankara", "TR42 (Konya)": "Konya",
                "TR07 (Antalya)": "Antalya", "TR67 (Zonguldak)": "Zonguldak",
                "TR25 (Erzurum)": "Erzurum"}
MAP_MUZE     = {c: c for c in CITIES}
MAP_OGRENCI  = {"İSTANBUL": "İstanbul", "İZMİR": "İzmir", "ANKARA": "Ankara",
                "KONYA": "Konya", "ANTALYA": "Antalya", "ZONGULDAK": "Zonguldak",
                "ERZURUM": "Erzurum"}


def cost_index():
    """Agirlikli maliyet endeksi (maliyett.xlsx)."""
    ws = load_workbook("maliyett.xlsx", data_only=True).active
    out = {}
    for r in ws.iter_rows(min_row=5, values_only=True):
        if (sehir := MAP_MALIYET.get(str(r[1]).strip() if r[1] else "")):
            try:
                out[sehir] = round(
                    float(r[5])*0.35 + float(r[3])*0.25 + float(r[7])*0.20
                    + float(r[6])*0.10 + float(r[8])*0.10, 2)
            except (TypeError, ValueError):
                pass
    return out


def culture_score():
    """Min-max normalize kultur skoru 1-5 (muze.xlsx)."""
    ws   = load_workbook("muze.xlsx", data_only=True).active
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0]]

    ref  = next(r for r in rows if str(r[0]).strip() == "Türkiye")
    totals = [(ref[1] or 0)+(ref[5] or 0), (ref[2] or 0)+(ref[6] or 0),
              (ref[3] or 0)+(ref[7] or 0)]

    raw = {MAP_MUZE[str(r[0]).strip()]:
           [((r[i] or 0)+(r[i+4] or 0)) / totals[j]
            for j, i in enumerate((1, 2, 3))]
           for r in rows if str(r[0]).strip() in MAP_MUZE}

    def mmn(key):
        vals = [raw[s][key] for s in raw]
        lo, hi = min(vals), max(vals)
        return {s: (raw[s][key]-lo)/(hi-lo) if hi != lo else 1.0 for s in raw}

    n = [mmn(k) for k in range(3)]
    W = [0.20, 0.35, 0.45]
    return {s: round(1 + sum(W[k]*n[k][s] for k in range(3)) * 4, 2) for s in raw}


def student_count():
    """Il bazinda toplam ogrenci sayisi (ogrenci.xlsx)."""
    ws, out = load_workbook("ogrenci.xlsx", data_only=True).active, {c: 0 for c in CITIES}
    for r in ws.iter_rows(min_row=5, values_only=True):
        if r[1] and str(r[1]).strip().upper() not in ("NAN", "TOPLAM"):
            if (sehir := MAP_OGRENCI.get(str(r[3] or "").strip().upper())):
                try:
                    out[sehir] += int(str(r[19]).replace(",", ""))
                except (TypeError, ValueError):
                    pass
    return out


def db_write(cost, culture, student):
   
    with psycopg2.connect(**DB) as conn, conn.cursor() as cur:
        cur.execute("SELECT id, city_name FROM cities;")
        existing = {name: cid for cid, name in cur.fetchall()}

        for s in sorted(CITIES):
            vals = (student.get(s, 0), culture.get(s), cost.get(s))
            if s in existing:
                cur.execute(
                    "UPDATE cities SET total_student_count=%s, culture_score=%s,"
                    " total_cost_index=%s WHERE id=%s;", (*vals, existing[s]))
            else:
                cur.execute(
                    "INSERT INTO cities (city_name,total_student_count,culture_score,"
                    "total_cost_index) VALUES (%s,%s,%s,%s);", (s, *vals))
            print(f"  {'UPDATE' if s in existing else 'INSERT':6}  {s}")

        cur.execute("SELECT id,city_name,total_cost_index,culture_score,"
                    "total_student_count FROM cities ORDER BY city_name;")
        print(f"\n  {'id':>3}  {'Sehir':<14}{'cost':>6}  {'culture':>8}  {'students':>10}")
        print("  " + "─" * 46)
        for r in cur.fetchall():
            print(f"  {r[0]:>3}  {r[1]:<14}{r[2]:>6}  {r[3]:>8}  {r[4]:>10,}")


if __name__ == "__main__":
    cost, culture, student = cost_index(), culture_score(), student_count()

    print(f"  {'Sehir':<14}{'cost':>6}  {'culture':>8}  {'students':>12}")
    print("  " + "─" * 44)
    for s in sorted(CITIES):
        print(f"  {s:<14}{cost.get(s,'?'):>6}  {culture.get(s,'?'):>8}  {student.get(s,0):>12,}")

    with open("hesaplanan_veriler.json", "w", encoding="utf-8") as f:
        json.dump({s: {"total_cost_index": cost.get(s), "culture_score": culture.get(s),
                       "total_student_count": student.get(s, 0)}
                   for s in sorted(CITIES)}, f, ensure_ascii=False, indent=2)

    print("\n  → hesaplanan_veriler.json kaydedildi\n")

    try:
        db_write(cost, culture, student)
        print("\nETL tamamlandi.")
    except psycopg2.Error as e:
        print(f"\n[DB HATA] {e}")
